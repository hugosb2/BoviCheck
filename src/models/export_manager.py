import json
import io
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

from .definitions import VERSION_NUMBER

def backup_to_json_string(indices_data: dict, selected_names: list[str]) -> str:
    to_backup = {name: indices_data[name] for name in selected_names if name in indices_data}
    
    backup_content = {
        "app_version": VERSION_NUMBER,
        "backup_timestamp": datetime.now().isoformat(),
        "selected_indices_data": to_backup
    }
    return json.dumps(backup_content, ensure_ascii=False, indent=4)

def restore_from_json_string(json_string: str) -> tuple[bool, str, dict]:
    try:
        data = json.loads(json_string)
        indices_to_restore = data.get("selected_indices_data")
        if not isinstance(indices_to_restore, dict):
            return False, "Erro: Formato de backup inválido.", {}
        
        for index_name, results in indices_to_restore.items():
            for i, result in enumerate(results):
                if "id" not in result or not result["id"]:
                    result["id"] = f"restored_{datetime.now().timestamp()}_{index_name}_{i}"
        
        count = len(indices_to_restore)
        msg = f"{count} índice(s) restaurado(s) com sucesso." if count > 0 else "Nenhum dado válido para restaurar."
        return True, msg, indices_to_restore
    except json.JSONDecodeError:
        return False, "Erro: Arquivo de backup não é um JSON válido.", {}
    except Exception as e:
        return False, f"Erro inesperado: {e}", {}

def generate_spreadsheet_bytes(indices_data: dict, selected_names: list[str]) -> bytes | None:
    if not OPENPYXL_AVAILABLE:
        return None
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dados dos Índices"
    header = ["Nome do Índice", "Índice (Valor e Unidade)", "Hora", "Data"]
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    indices_to_export = sorted([name for name in selected_names if name in indices_data])

    for name in indices_to_export:
        for result in indices_data.get(name, []):
            row = [name, result.get("Resultado", "N/A"), result.get("Hora", "N/A"), result.get("Data", "N/A")]
            sheet.append(row)

    for col_idx, column_cells in enumerate(sheet.columns, 1):
        max_len = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        sheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    
    with io.BytesIO() as excel_bytes_io:
        workbook.save(excel_bytes_io)
        return excel_bytes_io.getvalue()

def generate_pdf_bytes(indices_data: dict, selected_names: list[str]) -> bytes | None:
    if not FPDF_AVAILABLE:
        return None

    pdf = FPDF()
    pdf.add_page()
    
    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 10, "Relatório de Índices - BoviCheck", ln=True, align='C')
    pdf.ln(5)

    pdf.set_font("helvetica", "B", 10)
    col_widths = [80, 45, 25, 25]
    header = ["Nome do Índice", "Resultado", "Hora", "Data"]
    for i, item in enumerate(header):
        pdf.cell(col_widths[i], 8, item, border=1, align='C')
    pdf.ln()

    pdf.set_font("helvetica", "", 9)
    indices_to_export = sorted([name for name in selected_names if name in indices_data])

    for name in indices_to_export:
        for result in indices_data.get(name, []):
            row = [
                name,
                result.get("Resultado", "N/A"),
                result.get("Hora", "N/A"),
                result.get("Data", "N/A")
            ]
            y_before = pdf.get_y()
            pdf.multi_cell(col_widths[0], 6, row[0], border='L', align='L')
            x_after_multi_cell = pdf.get_x()
            y_after_multi_cell = pdf.get_y()
            
            pdf.set_xy(x_after_multi_cell + col_widths[0], y_before)
            
            pdf.cell(col_widths[1], (y_after_multi_cell - y_before), row[1], border=0)
            pdf.cell(col_widths[2], (y_after_multi_cell - y_before), row[2], border=0, align='C')
            pdf.cell(col_widths[3], (y_after_multi_cell - y_before), row[3], border='R', align='C')
            pdf.ln()

    pdf.cell(sum(col_widths), 0, '', 'T')
    pdf.ln(10)

    pdf.set_font("helvetica", "I", 8)
    pdf.set_text_color(128)
    pdf.cell(0, 10, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} com BoviCheck v{VERSION_NUMBER}", align='C')

    return pdf.output(dest='S').encode('latin-1')